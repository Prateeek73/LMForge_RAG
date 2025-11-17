from django.views import View
from django.http import StreamingHttpResponse, HttpRequest, HttpResponse
from urllib.parse import urlparse
import time
import re

from rest_framework.views import APIView
from rest_framework.request import Request
import logging
from typing import Dict, Any

from rest_framework.response import Response
from rest_framework import status
import requests
import json
from bs4 import BeautifulSoup
from django.shortcuts import render
from openpyxl import load_workbook
from io import BytesIO
from ..models.scraped_data import ScrapedData  # Import the model to save data
import pdfplumber
import markdown

from ..utils.content_extractor import extract_article_content
logger = logging.getLogger(__name__)
MAX_TITLE_LENGTH = 100  # Maximum length for ScrapedData title field
MAX_URL_TITLE_LENGTH = 95  # Maximum length when using URL as title (leave room for "scraped")


def remove_emojis(text):
    if not text:
        return ""
    # Remove characters outside the Basic Multilingual Plane (BMP)
    return re.sub(r'[\U00010000-\U0010FFFF]', '', text)

class ScrapeDataView(View):
    def stream_scrape_events(self, request):
        """
        A generator function that performs scraping and yields Server-Sent Events.
        (Structure from 'Temp')
        """
        url = request.GET.get('url')
        title = request.GET.get('title', '') 
        source_type = request.GET.get('source_type')

        def send_event(event_type, data):
            """Helper to format data as a Server-Sent Event."""
            json_data = json.dumps({"type": event_type, "data": data})
            return f"data: {json_data}\n\n"

        if not url:
            yield send_event('error', {'message': 'Please provide a URL.'})
            return

        try:
            yield send_event('progress', {'message': f"Connecting to {url}..."})

            content = None
            file_type = None

            if source_type == 'reddit':
                parsed_url = urlparse(url)
                headers = {'User-Agent': 'My-Django-Scraper-App/1.2'}
                
                if '/comments/' in parsed_url.path: # Handle specific post
                    api_url = url.rstrip('/') + ".json"
                    yield send_event('progress', {'message': f"Requesting data from Reddit API: {api_url}"})
                    response = requests.get(api_url, headers=headers)
                    response.raise_for_status()
                    data = response.json()

                    file_type = 'reddit_post'
                    yield send_event('progress', {'message': 'Parsing Reddit post and comments...'})
                    
                    post_data = data[0]['data']['children'][0]['data']
                    post_title = post_data.get('title', 'No Title')
                    post_author = post_data.get('author', 'Unknown Author')
                    post_text = post_data.get('selftext', 'No content text.')
                    
                    content_lines = [
                        f"Title: {post_title}",
                        f"Author: u/{post_author}",
                        "--- POST CONTENT ---",
                        post_text,
                        "\n--- COMMENTS ---"
                    ]

                    comments_data = data[1]['data']['children']
                    
                    content = "\n".join(content_lines)
                    # Use Reddit post title if user didn't provide one
                    if not title.strip():
                        title = post_title

                elif parsed_url.path.startswith('/r/'): # Handle subreddit
                    file_type = 'reddit_subreddit_full'
                    subreddit_name = parsed_url.path.split('/')[2]
                    base_api_url = f"https://www.reddit.com/r/{subreddit_name}"

                    content_lines = [f"Scraped Posts and Comments from r/{subreddit_name} (All Filters):\n" + "="*40]
                    
                    # Define the filters to scrape
                    simple_filters = ['hot', 'new', 'rising', 'best']
                    top_filters = [
                        ('hour', 'Now (Top)'), 
                        ('day', 'Today (Top)'), 
                        ('week', 'This Week (Top)'), 
                        ('month', 'This Month (Top)'), 
                        ('year', 'This Year (Top)'), 
                        ('all', 'All Time (Top)')
                    ]
                    
                    all_posts_to_scrape = [] # This will hold the 'children' data objects
                    seen_post_ids = set() # To avoid duplicates

                    # --- 1. Get Simple Filters (hot, new, rising, best) ---
                    for f in simple_filters:
                        filter_url = f"{base_api_url}/{f}.json?limit=25"
                        yield send_event('progress', {'message': f'Requesting /r/{subreddit_name} [{f}] listing...'})
                        try:
                            time.sleep(0.5) # Be nice to the API
                            response = requests.get(filter_url, headers=headers)
                            response.raise_for_status()
                            data = response.json()
                            new_posts = data.get('data', {}).get('children', [])
                            yield send_event('progress', {'message': f'Found {len(new_posts)} posts in [{f}].'})
                            
                            for post_item in new_posts:
                                post_id = post_item.get('data', {}).get('id')
                                if post_id and post_id not in seen_post_ids:
                                    all_posts_to_scrape.append(post_item)
                                    seen_post_ids.add(post_id)
                        except Exception as e:
                            yield send_event('error', {'message': f'Failed to get [{f}] listing: {str(e)}'})

                    # --- 2. Get Top Filters (all time ranges) ---
                    for t_param, t_name in top_filters:
                        filter_url = f"{base_api_url}/top.json?t={t_param}&limit=25"
                        yield send_event('progress', {'message': f'Requesting /r/{subreddit_name} [{t_name}] listing...'})
                        try:
                            time.sleep(0.5) # Be nice to the API
                            response = requests.get(filter_url, headers=headers)
                            response.raise_for_status()
                            data = response.json()
                            new_posts = data.get('data', {}).get('children', [])
                            yield send_event('progress', {'message': f'Found {len(new_posts)} posts in [{t_name}].'})

                            for post_item in new_posts:
                                post_id = post_item.get('data', {}).get('id')
                                if post_id and post_id not in seen_post_ids:
                                    all_posts_to_scrape.append(post_item)
                                    seen_post_ids.add(post_id)
                        except Exception as e:
                            yield send_event('error', {'message': f'Failed to get [{t_name}] listing: {str(e)}'})

                    # --- 3. Now scrape all the collected, unique posts ---
                    total_posts_to_scrape = len(all_posts_to_scrape)
                    yield send_event('progress', {'message': f'Collected {total_posts_to_scrape} unique posts across all filters. Now scraping content and comments...'})

                    for i, post_item in enumerate(all_posts_to_scrape):
                        post_data = post_item['data']
                        post_title = post_data.get('title', 'No Title')
                        permalink = post_data.get('permalink')

                        if not permalink: 
                            content_lines.append(f"\n[Skipping post with no permalink: {post_title}]")
                            continue
                        
                        yield send_event('progress', {'message': f'({i+1}/{total_posts_to_scrape}) Scraping: "{post_title[:50]}..."'})
                        
                        post_api_url = f"https://www.reddit.com{permalink.rstrip('/')}.json"
                        
                        content_lines.append(f"\n\n{'='*20}\nPOST: {post_title}\n{'='*20}")

                        try:
                            time.sleep(0.5) # Respect Reddit's rate limits
                            post_response = requests.get(post_api_url, headers=headers)
                            post_response.raise_for_status()
                            post_and_comment_data = post_response.json()
                            
                            post_content_data = post_and_comment_data[0]['data']['children'][0]['data']
                            post_text = post_content_data.get('selftext', '')
                            post_author = post_content_data.get('author', 'Unknown')
                            
                            content_lines.append(f"Author: u/{post_author}")
                            
                            if post_text:
                                content_lines.append(f"\n--- POST CONTENT ---\n{post_text}\n")
                            else:
                                content_lines.append(f"\n[No self-text for this post.]\n")

                            content_lines.append("--- COMMENTS ---")
                            comments_data = post_and_comment_data[1]['data']['children']
                            
                            comment_count = 0
                            def get_all_comments(comment_list, depth=0):
                                nonlocal comment_count
                                if not comment_list:
                                    return

                                for comment in comment_list:
                                    if comment.get('kind') != 't1': # t1 is a comment
                                        continue 
                                        
                                    if 'data' in comment and 'body' in comment['data']:
                                        comment_author = comment['data'].get('author', 'Unknown')
                                        comment_body = comment['data'].get('body', '')
                                        indent = "  " * depth
                                        content_lines.append(f"\n{indent}> u/{comment_author}:\n{indent}{comment_body.replace(chr(10), chr(10) + indent)}\n")
                                        comment_count += 1
                                        
                                        # Recurse for replies
                                        replies = comment['data'].get('replies')
                                        if replies and 'data' in replies and 'children' in replies['data']:
                                            get_all_comments(replies['data']['children'], depth + 1)
                            
                            get_all_comments(comments_data)
                            
                            if comment_count == 0:
                                content_lines.append("No comments found for this post.")

                        except Exception as post_e:
                            content_lines.append(f"\n[Could not fetch content for post '{post_title}'. Error: {str(post_e)}]")
                    
                    content = "\n".join(content_lines)
                    # Use subreddit name in title if user didn't provide one
                    if not title.strip():
                        title = f"Scrape of r/{subreddit_name}"

                else:
                    yield send_event('error', {'message': 'Invalid Reddit URL. Must be a post or a subreddit.'})
                    return
            
            else:
                yield send_event('progress', {'message': 'Scraping generic URL...'})
                response = requests.get(url)
                response.raise_for_status()
                content_type = response.headers.get('Content-Type', '').lower()

                if 'application/json' in content_type:
                    file_type = 'json'
                    scraped_content = json.dumps(response.json(), indent=4)
                elif 'application/xml' in content_type or 'text/xml' in content_type:
                    file_type = 'xml'
                    scraped_content = response.content.decode('utf-8')
                elif 'text/plain' in content_type:
                    file_type = 'text'
                    scraped_content = response.content.decode('utf-8')
                elif 'text/html' in content_type:
                    file_type = 'html'
                    soup = BeautifulSoup(response.content, 'html.parser')
                    for script in soup(["script", "style", "meta", "noscript"]):
                        script.extract()
                    main_content = soup.find("article")
                    if not main_content:
                        main_content = soup.find("div", {"class": "content"})
                    if main_content:
                        scraped_content = main_content.get_text(separator="\n", strip=True)
                    else:
                        scraped_content = soup.get_text(separator="\n", strip=True)
                    scraped_content = "\n".join([line.strip() for line in scraped_content.split("\n") if line.strip()])
                elif 'text/csv' in content_type or 'application/csv' in content_type:
                    file_type = 'csv'
                    scraped_content = response.content.decode('utf-8')
                elif 'application/vnd.ms-excel' in content_type or 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
                    file_type = 'xlsx'
                    binary_content = response.content
                else:
                    yield send_event('error', {'message': f'Unsupported content type: {content_type}'})
                    return

                content_type: str = response.headers.get("content-type", "").lower()

                if "application/json" in content_type or (response.text and response.text.strip().startswith("{")):
                    yield send_event('progress', {'message': 'Parsing JSON...'})
                    content = json.dumps(response.json(), indent=2)
                    file_type = "json"

                elif any(x in content_type for x in ("application/xml", "text/xml", "application/rss+xml")):
                    yield send_event('progress', {'message': 'Parsing XML/RSS...'})
                    content = response.text
                    file_type = "xml"

                elif "text/plain" in content_type:
                    yield send_event('progress', {'message': 'Parsing plain text...'})
                    content = response.text
                    file_type = "text"

                elif "text/csv" in content_type or url.lower().endswith(".csv"):
                    yield send_event('progress', {'message': 'Parsing CSV...'})
                    content = response.text
                    file_type = "csv"

                elif any(x in content_type for x in ("excel", "spreadsheetml", "vnd.openxmlformats")) or url.lower().endswith(".xlsx"):
                    yield send_event('progress', {'message': 'Parsing Excel (XLSX)...'})
                    try:
                        bio = BytesIO(response.content)
                        wb = load_workbook(filename=bio, read_only=True)
                        ws = wb[wb.sheetnames[0]]
                        rows = []
                        for row in ws.iter_rows(values_only=True):
                            rows.append(",".join([str(c) if c is not None else "" for c in row]))
                        content = "\n".join(rows)
                        file_type = "xlsx"
                    except Exception as e:
                        logger.exception("Failed to parse xlsx: %s", e)
                        yield send_event('error', {'message': "Failed to parse xlsx file"})
                        return
                
                else:
                    yield send_event('progress', {'message': 'Parsing HTML...'})
                    try:
                        result: Dict[str, Any] = extract_article_content(response.content, url)
                        content = result.get("body", "") or ""

                        if not content.strip():
                            logger.info("Extractor returned empty body for %s; falling back to simple HTML parsing", url)
                            soup = BeautifulSoup(response.content, "html.parser")
                            article = soup.find("article") or soup.find(class_="content") or soup.find("main")
                            if article:
                                text = article.get_text("\n\n", strip=True)
                            else:
                                body = soup.body
                                text = body.get_text("\n\n", strip=True) if body else soup.get_text("\n\n", strip=True)
                            content = text

                        if not title or not title.strip():
                            extracted_title = result.get("title") or ""
                            if extracted_title:
                                title = extracted_title 
                        file_type = "html"
                    except Exception as e:
                        logger.exception("Extractor failed, falling back to simple HTML parsing: %s", e)
                        soup = BeautifulSoup(response.content, "html.parser")
                        article = soup.find("article") or soup.find(class_="content") or soup.find("main")
                        if article:
                            text = article.get_text("\n\n", strip=True)
                        else:
                            body = soup.body
                            text = body.get_text("\n\n", strip=True) if body else soup.get_text("\n\n", strip=True)
                        content = text
                        file_type = "html"


            yield send_event('progress', {'message': 'Scraping complete. Saving to database...'})
            
            # Use emoji remover from 'Temp'
            cleaned_content = remove_emojis(content)

            # Use provided title, or extracted title, or fallback to URL
            final_title = title or (url[:MAX_URL_TITLE_LENGTH] if url else "scraped")
            
            ScrapedData.objects.create(
                url=url,
                file_type=file_type,
                content=cleaned_content,
                title=final_title[:MAX_TITLE_LENGTH] # Ensure max length
            )

            final_data = {
                'success': f'Successfully saved data from {url} to the database.',
                'url': url,
                'file_type': file_type,
                'content': cleaned_content
            }
            yield send_event('complete', final_data)
        
        except Exception as e:
            logger.exception("An unexpected error occurred during scrape: %s", e)
            yield send_event('error', {'message': f'An unexpected error occurred: {str(e)}'})
            return

    def get(self, request):
        """
        Returns a streaming response to send live scraping updates.
        (Kept from 'Temp')
        """
        response = StreamingHttpResponse(self.stream_scrape_events(request), content_type='text/event-stream')
        response['Cache-Control'] = 'no-cache'
        return response

class UploadPDFView(APIView):
    """Upload a PDF and convert it to text/html/json as requested."""

    def post(self, request: Request) -> Response:
        pdf_file = request.FILES.get("pdf_file")
        output_format = request.POST.get("output_format") or request.data.get("output_format") or "text"
        title = request.POST.get("title") or request.data.get("title") or (getattr(pdf_file, 'name', '') if pdf_file else "uploaded_pdf")

        if not pdf_file:
            return Response({"error": "No PDF file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            text_parts = []
            with pdfplumber.open(pdf_file) as pdf:
                for p in pdf.pages:
                    text_parts.append(p.extract_text() or "")
            text = "\n\n".join([t for t in text_parts if t])

            if output_format == "html":
                content = markdown.markdown(text)
                file_type = "html"
            elif output_format == "json":
                content = json.dumps({"text": text}, indent=2)
                file_type = "json"
            else: # Default to text
                content = text
                file_type = "text"

            scraped_record: ScrapedData = ScrapedData.objects.create(
                url="uploaded_pdf", # Use placeholder URL
                file_type=file_type,
                content=content,
                title=title[:MAX_TITLE_LENGTH]
            )
        except Exception as e:
            logger.exception("PDF conversion failed: %s", e)
            return Response({"error": "Failed to convert PDF"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Modified to return content, similar to 'Temp'
        return Response({
            "success": "PDF converted and saved", 
            "id": scraped_record.id, 
            "file_type": scraped_record.file_type, 
            "content": scraped_record.content
        }, status=status.HTTP_200_OK)

def scrape_view(request: HttpRequest) -> HttpResponse:
    """Render the scrape view with the latest scraped data."""
    latest: ScrapedData | None = ScrapedData.objects.order_by("-created_at").first()
    return render(request, "scrape.html", {"latest_scraped_data": latest})


class SaveManualTextView(APIView):
    """Save manually entered text to ScrapedData."""
    
    def post(self, request: Request) -> Response:
        text: str | None = request.data.get("text") or request.POST.get("text")
        title: str = request.data.get("title") or request.POST.get("title") or "manual"
        if not text:
            return Response({"error": "No text provided"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            scraped_record: ScrapedData = ScrapedData.objects.create(
                url="manual", # Use placeholder URL
                file_type="text",
                content=text,
                title=title[:MAX_TITLE_LENGTH]
            )
        except Exception as e:
            logger.exception("Failed to save manual text: %s", e)
            return Response({"error": "Failed to save text"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        return Response({
            "success": "Text saved", 
            "id": scraped_record.id, 
            "file_type": scraped_record.file_type,
            "content": scraped_record.content
        }, status=status.HTTP_200_OK)